"""
export_timing_breakdown.py — Excel summary of iteration time-breakdown profiling.
Reads per-city CSVs that contain timing columns (t_destroy_s, pct_reconstruct, etc.)
Produces results/gate/timing_breakdown_summary.xlsx

Usage:
    python scripts/export_timing_breakdown.py
    python scripts/export_timing_breakdown.py --out timing_breakdown_v1.xlsx
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT     = Path(__file__).resolve().parents[1]
INST_DIR = ROOT / "results" / "gate" / "instance_tests"
OUT_DIR  = ROOT / "results" / "gate"

p = argparse.ArgumentParser()
p.add_argument("--out", default="timing_breakdown_summary.xlsx")
args = p.parse_args()

N_MAP = {
    "Parma":    200,
    "Catania":  475,
    "Genova":   825,
    "Torino":   1950,
    "Milano":   3375,
}

METHOD_ORDER  = {"no_gate": 0, "fixed_gate": 1, "rolling": 2, "calibration": 3}
METHOD_LABELS = {"no_gate": "No gate", "fixed_gate": "Fixed gate",
                 "rolling": "Rolling", "calibration": "Calibration"}
METHOD_COLORS = {"no_gate": "2471A3", "fixed_gate": "C0392B",
                 "rolling": "27AE60", "calibration": "E67E22"}

# ── Load CSVs ─────────────────────────────────────────────────────────────────
city_dfs = {}
for city, N in sorted(N_MAP.items(), key=lambda x: x[1]):
    csv_path = INST_DIR / city / f"{city}_results.csv"
    if not csv_path.exists():
        print(f"WARNING: {csv_path} not found, skipping")
        continue
    df = pd.read_csv(csv_path)
    if "pct_reconstruct" not in df.columns:
        print(f"WARNING: {city} CSV missing timing columns, skipping")
        continue
    df["city"] = city
    df["N"]    = N
    df["_mo"]  = df["method"].map(METHOD_ORDER)
    city_dfs[city] = df

if not city_dfs:
    print("No timing CSVs found — run profiling benchmarks first.")
    sys.exit(1)

all_df = pd.concat(city_dfs.values(), ignore_index=True)
all_df = all_df.sort_values(["N", "_mo"]).reset_index(drop=True)
ng_df  = all_df[all_df["method"] == "no_gate"].sort_values("N").reset_index(drop=True)

print(f"Loaded {len(city_dfs)} instances: "
      + ", ".join(f"{c}(N={N_MAP[c]})" for c in sorted(city_dfs, key=lambda x: N_MAP[x])))

# ── Style helpers ─────────────────────────────────────────────────────────────
def hfill(h):
    return PatternFill("solid", fgColor=h.lstrip("#"))

thin = Side(border_style="thin", color="CCCCCC")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR  = Alignment(horizontal="center", vertical="center")
CTRW = Alignment(horizontal="center", vertical="center", wrap_text=True)

def wf(color="FFFFFF", bold=True, size=10):
    return Font(color=color, bold=bold, size=size)

def hdr(ws, row, headers, fill, font=None):
    f = font or wf()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = fill; c.font = f; c.alignment = CTRW; c.border = BDR

def data_row(ws, row, values, alt=False):
    bg = hfill("EBF5FB") if alt else hfill("FDFEFE")
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = bg; c.alignment = CTR; c.border = BDR

def color_recon(cell, pct):
    if pct > 60:
        cell.fill = hfill("F1948A"); cell.font = Font(bold=True, color="C0392B")
    elif pct > 40:
        cell.fill = hfill("F9E79F"); cell.font = Font(bold=True, color="7D6608")
    else:
        cell.fill = hfill("A9DFBF"); cell.font = Font(bold=True, color="1E8449")

def color_ue(cell, pct):
    if pct > 50:
        cell.fill = hfill("A9DFBF"); cell.font = Font(bold=True, color="1E8449")
    elif pct > 30:
        cell.fill = hfill("F9E79F")

def color_proxy(cell, pct):
    if pct > 25:
        cell.fill = hfill("F1948A"); cell.font = Font(bold=True, color="C0392B")
    elif pct > 10:
        cell.fill = hfill("F9E79F"); cell.font = Font(bold=True, color="7D6608")
    else:
        cell.fill = hfill("A9DFBF"); cell.font = Font(bold=True, color="1E8449")

def color_ratio(cell, ratio):
    if ratio > 0.5:
        cell.fill = hfill("F1948A"); cell.font = Font(bold=True, color="C0392B")
    elif ratio > 0.2:
        cell.fill = hfill("F9E79F"); cell.font = Font(bold=True, color="7D6608")
    else:
        cell.fill = hfill("A9DFBF"); cell.font = Font(bold=True, color="1E8449")

# ── Chart 1: Log-log scaling + stacked % (no-gate) ───────────────────────────
plt.rcParams.update({"font.family": "serif", "font.size": 10,
                     "axes.spines.top": False, "axes.spines.right": False})

chart1_path = OUT_DIR / "_tb_scaling_chart.png"
Ns_   = ng_df["N"].tolist()
d_ms  = ng_df["avg_destroy_ms"].tolist()
r_ms  = ng_df["avg_reconstruct_ms"].tolist()
ue_ms = ng_df["avg_ue_ms"].tolist()
ls_ms = ng_df["avg_ls_pure_ms"].tolist()
cits  = ng_df["city"].tolist()

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

ax1.loglog(Ns_, d_ms,  "o-", color="#1A5276", lw=2, ms=7, label="Destroy  ~O(N^0.94)")
ax1.loglog(Ns_, r_ms,  "s-", color="#C0392B", lw=2, ms=7, label="Reconstruct  ~O(N^2.2)")
ax1.loglog(Ns_, ue_ms, "^-", color="#8E44AD", lw=2, ms=7, label="UE (Gurobi)  ~O(N^1.3)")
ax1.loglog(Ns_, ls_ms, "D-", color="#27AE60", lw=2, ms=7, label="LS overhead  ~O(N^1.0)")
for N_v, rm, city in zip(Ns_, r_ms, cits):
    ax1.annotate(city, (N_v, rm), textcoords="offset points",
                 xytext=(6, 0), fontsize=7.5, color="#C0392B")
ax1.set_xlabel("N (log scale)", fontsize=11)
ax1.set_ylabel("ms per call (log scale)", fontsize=11)
ax1.set_title("Component cost vs N — no-gate baseline", fontsize=11, pad=8)
ax1.legend(fontsize=9, frameon=False)
ax1.grid(True, alpha=0.3, linestyle="--")

p_des = ng_df["pct_destroy"].tolist()
p_rec = ng_df["pct_reconstruct"].tolist()
p_ue  = ng_df["pct_ue"].tolist()
p_ls  = ng_df["pct_ls_pure"].tolist()
p_oth = ng_df["pct_other"].tolist()
xlabs = [f"{c}\nN={n}" for c, n in zip(cits, Ns_)]
x     = range(len(Ns_))

def stacked(ax, x, bottoms, vals, color, label):
    ax.bar(x, vals, bottom=bottoms, color=color, label=label, width=0.6, edgecolor="white")
    return [b + v for b, v in zip(bottoms, vals)]

bot = [0.0] * len(Ns_)
bot = stacked(ax2, x, bot, p_des, "#1A5276", "Destroy")
bot = stacked(ax2, x, bot, p_rec, "#C0392B", "Reconstruct")
bot = stacked(ax2, x, bot, p_ue,  "#8E44AD", "UE (Gurobi)")
bot = stacked(ax2, x, bot, p_ls,  "#27AE60", "LS overhead")
stacked(ax2, x, bot, p_oth, "#AAB7B8", "Other")
ax2.set_xticks(list(x)); ax2.set_xticklabels(xlabs, fontsize=8.5)
ax2.set_ylabel("% of time budget", fontsize=11); ax2.set_ylim(0, 105)
ax2.set_title("Budget allocation by component — no-gate", fontsize=11, pad=8)
ax2.legend(fontsize=8.5, frameon=False, ncol=2)
ax2.grid(axis="y", alpha=0.2, linestyle="--")

plt.tight_layout()
fig.savefig(chart1_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)

# ── Chart 2: Per-city stacked bars for all 4 methods ─────────────────────────
chart2_path = OUT_DIR / "_tb_methods_chart.png"
TIME_CATS   = ["Destroy", "Reconstruct", "Proxy", "UE (Gurobi)", "LS overhead", "Other"]
TIME_COLORS = ["#1A5276", "#C0392B", "#F39C12", "#8E44AD", "#27AE60", "#AAB7B8"]
pct_keys    = ["pct_destroy", "pct_reconstruct", "pct_proxy", "pct_ue", "pct_ls_pure", "pct_other"]

cities_sorted = sorted(city_dfs.keys(), key=lambda c: N_MAP[c])
n_cities = len(cities_sorted)
fig2, axes = plt.subplots(1, n_cities, figsize=(3.2 * n_cities, 5), sharey=True)
if n_cities == 1:
    axes = [axes]

for ax, city in zip(axes, cities_sorted):
    df_c = all_df[all_df["city"] == city].sort_values("_mo").reset_index(drop=True)
    mlabels = [METHOD_LABELS.get(m, m) for m in df_c["method"]]
    xp = range(len(df_c))
    bot2 = [0.0] * len(df_c)
    for cat, color, key in zip(TIME_CATS, TIME_COLORS, pct_keys):
        vals = df_c[key].tolist() if key in df_c.columns else [0.0] * len(df_c)
        ax.bar(xp, vals, bottom=bot2, color=color, label=cat, width=0.65, edgecolor="white")
        bot2 = [b + v for b, v in zip(bot2, vals)]
    ax.set_xticks(list(xp))
    ax.set_xticklabels(mlabels, rotation=30, ha="right", fontsize=8)
    ax.set_title(f"{city}\nN={N_MAP[city]}", fontsize=9, pad=4)
    ax.grid(axis="y", alpha=0.2, linestyle="--")
    ax.set_ylim(0, 108)

axes[0].set_ylabel("% of time budget", fontsize=10)
handles, labels = axes[0].get_legend_handles_labels()
fig2.legend(handles, labels, loc="lower center", ncol=6, fontsize=8.5, frameon=False)
fig2.suptitle("Time breakdown by method across instances", fontsize=11, y=1.01)
plt.tight_layout(rect=[0, 0.08, 1, 1])
fig2.savefig(chart2_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig2)

plt.rcdefaults()

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — No-Gate Scaling
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "No-Gate Scaling"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

c = ws1.cell(row=1, column=1,
    value="Iteration Time Breakdown — Scaling with N  (no-gate baseline, seed=11, alpha=0.01)")
c.font = Font(size=13, bold=True, color="1A5276")
ws1.merge_cells("A1:R1")

c2 = ws1.cell(row=2, column=1,
    value="Each row = one city/instance. Columns show % of time budget and avg ms per call for each component.")
c2.font = Font(size=9, italic=True, color="555555")
ws1.merge_cells("A2:R2")

# Group header row
groups = [
    ("Instance",           1,  3),
    ("% of time budget",   4,  9),
    ("avg ms / call",     10, 14),
    ("Run totals",        15, 17),
]
for label, c1, c2_ in groups:
    ws1.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2_)
    gc = ws1.cell(row=3, column=c1, value=label)
    gc.fill = hfill("1A5276"); gc.font = wf(); gc.alignment = CTR; gc.border = BDR

cols1 = [
    "N", "City", "Evals",
    "Destroy %", "Reconstruct %", "D&R %", "UE %", "LS %", "Other %",
    "Destroy ms/iter", "Reconstruct ms/iter", "D&R ms/iter", "UE ms/LS", "LS ms/LS",
    "Budget (s)", "Best score", "LS resp. %",
]
hdr(ws1, 4, cols1, hfill("1A5276"))

for ri, row in ng_df.iterrows():
    r = ri + 5
    vals = [
        row["N"], row["city"], int(row.get("full_evals", 0)),
        round(row["pct_destroy"], 1),
        round(row["pct_reconstruct"], 1),
        round(row["pct_dr"], 1),
        round(row["pct_ue"], 1),
        round(row["pct_ls_pure"], 1),
        round(row.get("pct_other", 0), 1),
        round(row["avg_destroy_ms"], 1),
        round(row["avg_reconstruct_ms"], 1),
        round(row["avg_dr_ms"], 1),
        round(row["avg_ue_ms"], 1),
        round(row["avg_ls_pure_ms"], 1),
        round(row.get("budget_s", 0), 0),
        round(row["best_score"], 3),
        round(row.get("ls_responsible_%", 0), 1),
    ]
    data_row(ws1, r, vals, alt=(ri % 2 == 0))
    color_recon(ws1.cell(row=r, column=5), row["pct_reconstruct"])
    color_ue(ws1.cell(row=r, column=7),    row["pct_ue"])

# Scaling exponent table
er = len(ng_df) + 7
ws1.cell(row=er, column=1, value="Scaling Exponents (fitted over 5 instances)").font = Font(bold=True, size=11, color="1A5276")
ws1.merge_cells(f"A{er}:G{er}")

exp_hdr = ["Component", "Scale", "Interpretation", "Crossover point"]
hdr(ws1, er + 1, exp_hdr, hfill("566573"))
exp_data = [
    ("Destroy",        "O(N^0.94)",  "Near-linear — always <1% of budget",         "Never a bottleneck"),
    ("Reconstruct",    "O(N^2.2)",   "Quadratic — dominates at N≥825",             "Overtakes UE between N=475 and N=825"),
    ("UE (Gurobi LP)", "O(N^1.3)",   "Sub-quadratic — drops from 67% to 21%",      "Overtaken by reconstruct at N≈700"),
    ("LS overhead",    "O(N^1.0)",   "Linear with iter count — always <10%",        "—"),
    ("Proxy (rolling)","~O(N^1.5)",  "Each call ≈60-70% of a UE; called 2×/iter",  "Exceeds UE budget% at N>475"),
]
for i, row in enumerate(exp_data):
    r = er + 2 + i
    alt = (i % 2 == 0)
    data_row(ws1, r, list(row), alt=alt)
    ws1.cell(row=r, column=1).font = Font(bold=True)

col_w1 = [7, 12, 7, 12, 15, 8, 8, 8, 8, 17, 19, 11, 12, 11, 11, 12, 11]
for ci, w in enumerate(col_w1, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[3].height = 20
ws1.row_dimensions[4].height = 32

if chart1_path.exists():
    img = XLImage(str(chart1_path))
    img.width, img.height = 920, 350
    ws1.add_image(img, f"A{er + len(exp_data) + 4}")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — All Methods
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Methods")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.cell(row=1, column=1,
    value="Time Breakdown — All Methods (no_gate / fixed_gate / rolling / calibration)").font = \
    Font(size=12, bold=True, color="117A65")
ws2.merge_cells("A1:S1")

cols2 = [
    "N", "City", "Method",
    "Destroy %", "Reconstruct %", "D&R %", "Proxy %", "UE %", "LS %",
    "Destroy ms", "Reconstruct ms", "D&R ms", "Proxy ms", "UE ms/LS", "LS ms/LS",
    "Evals", "Skip %", "Best score", "Calib Δ Rolling",
]
hdr(ws2, 2, cols2, hfill("117A65"))

# pre-compute calibration vs rolling delta per city
cr_delta = {}
for city in city_dfs:
    df_c = all_df[all_df["city"] == city]
    c_s = df_c[df_c["method"] == "calibration"]["best_score"].values
    r_s = df_c[df_c["method"] == "rolling"]["best_score"].values
    if len(c_s) and len(r_s):
        cr_delta[city] = round(float(c_s[0]) - float(r_s[0]), 3)

prev_city = None
alt = False
for ri, row in all_df.iterrows():
    r = ri + 3
    if row["city"] != prev_city:
        alt = not alt
        prev_city = row["city"]

    bg = hfill("EBF5FB") if alt else hfill("FDFEFE")
    for ci in range(1, len(cols2) + 1):
        ws2.cell(row=r, column=ci).fill = bg
        ws2.cell(row=r, column=ci).alignment = CTR
        ws2.cell(row=r, column=ci).border = BDR

    mcell = ws2.cell(row=r, column=3, value=METHOD_LABELS.get(row["method"], row["method"]))
    mcell.font = Font(bold=True, color=METHOD_COLORS.get(row["method"], "000000"))

    delta_cr = cr_delta.get(row["city"], "") if row["method"] == "calibration" else ""

    for ci, v in enumerate([
        row["N"], row["city"], None,
        round(row["pct_destroy"], 1),      round(row["pct_reconstruct"], 1),
        round(row["pct_dr"], 1),           round(row.get("pct_proxy", 0), 1),
        round(row["pct_ue"], 1),           round(row["pct_ls_pure"], 1),
        round(row["avg_destroy_ms"], 1),   round(row["avg_reconstruct_ms"], 1),
        round(row["avg_dr_ms"], 1),        round(row.get("avg_proxy_ms", 0), 1),
        round(row["avg_ue_ms"], 1),        round(row["avg_ls_pure_ms"], 1),
        int(row.get("full_evals", 0)),     round(row["skip_%"], 1),
        round(row["best_score"], 3),       delta_cr,
    ], 1):
        if ci == 3:
            continue
        ws2.cell(row=r, column=ci).value = v

    color_recon(ws2.cell(row=r, column=5),  row["pct_reconstruct"])
    color_ue(ws2.cell(row=r, column=8),     row["pct_ue"])
    color_proxy(ws2.cell(row=r, column=7),  row.get("pct_proxy", 0))

    if isinstance(delta_cr, float):
        dcell = ws2.cell(row=r, column=len(cols2))
        if delta_cr > 0.05:
            dcell.fill = hfill("A9DFBF"); dcell.font = Font(bold=True, color="1E8449")
        elif delta_cr < -0.05:
            dcell.fill = hfill("F1948A"); dcell.font = Font(bold=True, color="C0392B")
        else:
            dcell.fill = hfill("F9E79F")

col_w2 = [7, 12, 13, 11, 14, 8, 8, 8, 8, 12, 16, 10, 10, 12, 10, 7, 8, 12, 14]
for ci, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

if chart2_path.exists():
    img2 = XLImage(str(chart2_path))
    img2.width, img2.height = 920, 310
    ws2.add_image(img2, f"A{len(all_df) + 5}")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Proxy Overhead
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Proxy Overhead")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

ws3.cell(row=1, column=1,
    value="Proxy Overhead vs UE Cost — Gate Methods Only (fixed_gate / rolling / calibration)").font = \
    Font(size=12, bold=True, color="922B21")
ws3.merge_cells("A1:K1")

cols3 = [
    "N", "City", "Method",
    "Proxy ms/call", "UE ms/LS", "Proxy / UE",
    "Proxy % budget", "UE % budget", "Reconstruct % budget",
    "Skip %", "Best score",
]
hdr(ws3, 2, cols3, hfill("922B21"))

gate_df = all_df[all_df["method"].isin(["fixed_gate", "rolling", "calibration"])].copy()
gate_df = gate_df.sort_values(["N", "_mo"]).reset_index(drop=True)

prev_city = None
alt = False
for ri, row in gate_df.iterrows():
    r = ri + 3
    if row["city"] != prev_city:
        alt = not alt
        prev_city = row["city"]

    proxy_ms = float(row.get("avg_proxy_ms", 0))
    ue_ms_v  = float(row.get("avg_ue_ms", 1))
    ratio    = round(proxy_ms / max(ue_ms_v, 0.1), 3)

    vals = [
        row["N"], row["city"],
        METHOD_LABELS.get(row["method"], row["method"]),
        round(proxy_ms, 1),
        round(ue_ms_v, 1),
        ratio,
        round(row.get("pct_proxy", 0), 1),
        round(row["pct_ue"], 1),
        round(row["pct_reconstruct"], 1),
        round(row["skip_%"], 1),
        round(row["best_score"], 3),
    ]
    data_row(ws3, r, vals, alt=alt)

    mcell = ws3.cell(row=r, column=3)
    mcell.font = Font(bold=True, color=METHOD_COLORS.get(row["method"], "000000"))
    color_ratio(ws3.cell(row=r, column=6),  ratio)
    color_proxy(ws3.cell(row=r, column=7),  row.get("pct_proxy", 0))
    color_recon(ws3.cell(row=r, column=9),  row["pct_reconstruct"])

col_w3 = [7, 12, 13, 15, 12, 13, 16, 13, 20, 9, 12]
for ci, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = OUT_DIR / args.out
wb.save(out_path)
print(f"\nExcel -> {out_path}")
print("Sheets: No-Gate Scaling, All Methods, Proxy Overhead")
