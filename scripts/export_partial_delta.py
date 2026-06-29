"""
export_partial_delta.py — Excel summary of partial-delta reconstruction results.
Reads *_partial_delta_results.csv from results/gate/partial_delta/
Produces results/gate/partial_delta_summary.xlsx

Usage:
    python scripts/export_partial_delta.py
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT    = Path(__file__).resolve().parents[1]
RES_DIR = ROOT / "results" / "gate" / "partial_delta"
OUT     = ROOT / "results" / "gate" / "partial_delta_summary.xlsx"

# ── Load CSVs ─────────────────────────────────────────────────────────────────
frames = []
for csv_path in sorted(RES_DIR.glob("*_partial_delta_results.csv")):
    frames.append(pd.read_csv(csv_path))

if not frames:
    print("No CSV files found in", RES_DIR)
    sys.exit(1)

df = pd.concat(frames, ignore_index=True)
df = df.sort_values(["N", "refresh_R"], ascending=[True, False]).reset_index(drop=True)

# Add derived columns
cities_sorted = df.groupby("city")["N"].first().sort_values()
for city in cities_sorted.index:
    mask     = df["city"] == city
    baseline = df.loc[mask & (df["label"] == "baseline"), "best_score"].values[0]
    base_ms  = df.loc[mask & (df["label"] == "baseline"), "avg_recon_ms"].values[0]
    df.loc[mask, "delta_vs_baseline"] = (df.loc[mask, "best_score"] - baseline).round(3)
    df.loc[mask, "speedup"]           = (base_ms / df.loc[mask, "avg_recon_ms"].clip(0.01)).round(2)

# ── Style helpers ─────────────────────────────────────────────────────────────
def hf(h): return PatternFill("solid", fgColor=h.lstrip("#"))
thin = Side(border_style="thin", color="CCCCCC")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR  = Alignment(horizontal="center", vertical="center")
CTRW = Alignment(horizontal="center", vertical="center", wrap_text=True)
def wf(c="FFFFFF", b=True, sz=10): return Font(color=c, bold=b, size=sz)

METHOD_COLOR = {"baseline": "2471A3", "R=20%": "1E8449",
                "R=10%": "27AE60", "R=5%": "F39C12"}

def hdr(ws, row, headers, fill):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = fill; c.font = wf(); c.alignment = CTRW; c.border = BDR

def drow(ws, row, vals, alt=False):
    bg = hf("EBF5FB") if alt else hf("FDFEFE")
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = bg; c.alignment = CTR; c.border = BDR

def color_speedup(cell, v):
    if v >= 8:   cell.fill = hf("A9DFBF"); cell.font = Font(bold=True, color="1E8449")
    elif v >= 4: cell.fill = hf("F9E79F"); cell.font = Font(bold=True, color="7D6608")

def color_delta(cell, v):
    if v > 0.5:   cell.fill = hf("A9DFBF"); cell.font = Font(bold=True, color="1E8449")
    elif v > 0:   cell.fill = hf("D5F5E3")
    elif v < -1:  cell.fill = hf("F1948A"); cell.font = Font(bold=True, color="C0392B")

# ── Charts ────────────────────────────────────────────────────────────────────
plt.rcParams.update({"font.family": "serif", "font.size": 10,
                     "axes.spines.top": False, "axes.spines.right": False})

chart1_path = RES_DIR / "_pd_speedup_chart.png"
chart2_path = RES_DIR / "_pd_quality_chart.png"

cities  = list(cities_sorted.index)
configs = ["R=5%", "R=10%", "R=20%"]
colors  = {"R=5%": "#F39C12", "R=10%": "#27AE60", "R=20%": "#2471A3"}
x       = np.arange(len(cities))
width   = 0.25

# Chart 1: Speedup per city × config
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))

for ki, cfg in enumerate(configs):
    speedups = []
    for city in cities:
        row = df[(df["city"] == city) & (df["label"] == cfg)]
        speedups.append(float(row["speedup"].values[0]) if len(row) else 0)
    bars = ax1.bar(x + ki * width, speedups, width, label=cfg,
                   color=colors[cfg], edgecolor="white")
    for bar, v in zip(bars, speedups):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
                 f"{v:.1f}×", ha="center", fontsize=8, fontweight="bold")

ax1.set_xticks(x + width)
ax1.set_xticklabels([f"{c}\nN={cities_sorted[c]}" for c in cities], fontsize=9)
ax1.set_ylabel("Reconstruct speedup vs baseline", fontsize=10)
ax1.set_title("Partial-delta speedup by instance & R", fontsize=11, pad=8)
ax1.legend(fontsize=9, frameon=False)
ax1.axhline(1, color="#999", lw=1, ls="--")
ax1.grid(axis="y", alpha=0.2, ls="--")

# Chart 2: Quality delta per city × config
for ki, cfg in enumerate(configs):
    deltas = []
    for city in cities:
        row = df[(df["city"] == city) & (df["label"] == cfg)]
        deltas.append(float(row["delta_vs_baseline"].values[0]) if len(row) else 0)
    bars = ax2.bar(x + ki * width, deltas, width, label=cfg,
                   color=colors[cfg], edgecolor="white")
    for bar, v in zip(bars, deltas):
        yoff = 0.1 if v >= 0 else -0.4
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + yoff,
                 f"{v:+.1f}", ha="center", fontsize=8, fontweight="bold")

ax2.set_xticks(x + width)
ax2.set_xticklabels([f"{c}\nN={cities_sorted[c]}" for c in cities], fontsize=9)
ax2.set_ylabel("Score Δ vs full-recompute baseline", fontsize=10)
ax2.set_title("Solution quality: partial-delta vs baseline", fontsize=11, pad=8)
ax2.legend(fontsize=9, frameon=False)
ax2.axhline(0, color="#555", lw=1.2)
ax2.grid(axis="y", alpha=0.2, ls="--")

plt.tight_layout()
fig.savefig(chart1_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)

# Chart 3: Reconstruct ms vs N (log-log scaling)
chart3_path = RES_DIR / "_pd_scaling_chart.png"
fig3, ax3 = plt.subplots(figsize=(8, 5))
all_configs = ["baseline"] + configs
cfg_colors  = {"baseline": "#C0392B", "R=20%": "#2471A3",
               "R=10%": "#27AE60", "R=5%": "#F39C12"}
Ns_all = [cities_sorted[c] for c in cities]
for cfg in all_configs:
    ms_vals = []
    for city in cities:
        row = df[(df["city"] == city) & (df["label"] == cfg)]
        ms_vals.append(float(row["avg_recon_ms"].values[0]) if len(row) else np.nan)
    ax3.loglog(Ns_all, ms_vals, "o-", color=cfg_colors[cfg], lw=2, ms=7, label=cfg)
ax3.set_xlabel("N (log scale)", fontsize=11)
ax3.set_ylabel("Avg reconstruct ms/iter (log scale)", fontsize=11)
ax3.set_title("Reconstruct cost scaling — baseline vs partial-delta", fontsize=11, pad=8)
ax3.legend(fontsize=9, frameon=False)
ax3.grid(True, alpha=0.3, ls="--")
plt.tight_layout()
fig3.savefig(chart3_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig3)

plt.rcdefaults()

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Summary (one row per city × config)
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

ws1.cell(row=1, column=1,
    value="Partial-Delta Reconstruction — Between-Iteration Cache  (seed=11, no local search)").font = \
    Font(size=13, bold=True, color="1A5276")
ws1.merge_cells("A1:M1")

cols1 = [
    "N", "City", "Config", "R",
    "Iterations", "Best score", "Δ vs baseline",
    "Recon ms/iter", "Speedup",
    "UE ms/eval", "Recon %", "UE %",
    "Recon time (s)",
]
hdr(ws1, 2, cols1, hf("1A5276"))

prev_city = None
alt = False
for ri, row in df.iterrows():
    r = ri + 3
    if row["city"] != prev_city:
        alt = not alt
        prev_city = row["city"]

    bg = hf("EBF5FB") if alt else hf("FDFEFE")
    for ci in range(1, len(cols1) + 1):
        ws1.cell(row=r, column=ci).fill = bg
        ws1.cell(row=r, column=ci).alignment = CTR
        ws1.cell(row=r, column=ci).border = BDR

    lbl = row["label"]
    mcell = ws1.cell(row=r, column=3, value=lbl)
    mcell.font = Font(bold=True, color=METHOD_COLOR.get(lbl, "000000"))

    vals = [
        row["N"], row["city"], None,
        "" if row["refresh_R"] >= 1.0 else f"{int(row['refresh_R']*100)}%",
        int(row["iters"]),
        round(row["best_score"], 3),
        round(row["delta_vs_baseline"], 3),
        round(row["avg_recon_ms"], 1),
        round(row["speedup"], 2),
        round(row["avg_ue_ms"], 1),
        round(row["pct_reconstruct"], 1),
        round(row["pct_ue"], 1),
        round(row["t_reconstruct_s"], 1),
    ]
    for ci, v in enumerate(vals, 1):
        if ci == 3: continue
        ws1.cell(row=r, column=ci).value = v

    color_speedup(ws1.cell(row=r, column=9), row["speedup"])
    color_delta(ws1.cell(row=r, column=7), row["delta_vs_baseline"])

col_w1 = [7, 12, 12, 6, 10, 12, 14, 13, 10, 12, 10, 8, 14]
for ci, w in enumerate(col_w1, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# Embed charts
img_row = len(df) + 5
if chart1_path.exists():
    img = XLImage(str(chart1_path))
    img.width, img.height = 860, 340
    ws1.add_image(img, f"A{img_row}")

if chart3_path.exists():
    img3 = XLImage(str(chart3_path))
    img3.width, img3.height = 530, 340
    ws1.add_image(img3, f"K{img_row}")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Per-city breakdowns (one tab per city)
# ═══════════════════════════════════════════════════════════════════════════════
for city in cities:
    ws = wb.create_sheet(city)
    ws.sheet_view.showGridLines = False

    N_val = int(cities_sorted[city])
    ws.cell(row=1, column=1,
        value=f"{city}  (N={N_val})  —  Partial-Delta Reconstruction Results").font = \
        Font(size=12, bold=True, color="1A5276")
    ws.merge_cells("A1:K1")

    cols_c = [
        "Config", "R",
        "Iterations", "Best score", "Δ vs baseline",
        "Recon ms/iter", "Speedup",
        "UE ms/eval", "Recon %", "UE %",
        "Recon time (s)",
    ]
    hdr(ws, 2, cols_c, hf("117A65"))

    city_df = df[df["city"] == city].sort_values("refresh_R", ascending=False).reset_index(drop=True)
    baseline_score = city_df[city_df["label"] == "baseline"]["best_score"].values[0]

    for ri, row in city_df.iterrows():
        r = ri + 3
        bg = hf("EBF5FB") if ri % 2 == 0 else hf("FDFEFE")
        for ci in range(1, len(cols_c) + 1):
            ws.cell(row=r, column=ci).fill = bg
            ws.cell(row=r, column=ci).alignment = CTR
            ws.cell(row=r, column=ci).border = BDR

        lbl = row["label"]
        mcell = ws.cell(row=r, column=1, value=lbl)
        mcell.font = Font(bold=True, color=METHOD_COLOR.get(lbl, "000000"))

        vals_c = [
            None,
            "" if row["refresh_R"] >= 1.0 else f"{int(row['refresh_R']*100)}%",
            int(row["iters"]),
            round(row["best_score"], 3),
            round(row["delta_vs_baseline"], 3),
            round(row["avg_recon_ms"], 1),
            round(row["speedup"], 2),
            round(row["avg_ue_ms"], 1),
            round(row["pct_reconstruct"], 1),
            round(row["pct_ue"], 1),
            round(row["t_reconstruct_s"], 1),
        ]
        for ci, v in enumerate(vals_c, 1):
            if ci == 1: continue
            ws.cell(row=r, column=ci).value = v

        color_speedup(ws.cell(row=r, column=7), row["speedup"])
        color_delta(ws.cell(row=r, column=5), row["delta_vs_baseline"])

    col_wc = [13, 6, 10, 12, 14, 13, 10, 12, 10, 8, 14]
    for ci, w in enumerate(col_wc, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
cities_done = ", ".join(f"{c}(N={int(cities_sorted[c])})" for c in cities)
print(f"\nExcel -> {OUT}")
print(f"Instances: {cities_done}")
print(f"Rows: {len(df)}")
