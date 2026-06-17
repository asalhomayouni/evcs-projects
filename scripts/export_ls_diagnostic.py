"""
export_ls_diagnostic.py — Excel export for LS diagnostic benchmark results.
Reads all *_results.csv from results/gate/instance_tests/*/
Handles both old column format (new_best_via_ls_%) and new format (ls_responsible_%).

Usage:
    python scripts/export_ls_diagnostic.py
    python scripts/export_ls_diagnostic.py --out gate_ls_diagnostic_v2.xlsx
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT     = Path(__file__).resolve().parents[1]
INST_DIR = ROOT / "results" / "gate" / "instance_tests"
OUT_DIR  = ROOT / "results" / "gate"

p = argparse.ArgumentParser()
p.add_argument("--out", default="gate_ls_diagnostic_v1.xlsx")
args = p.parse_args()

# ── N lookup from filename ────────────────────────────────────────────────────
N_MAP = {
    "Reggio":      100,
    "Brescia":     175,
    "Trieste":     225,
    "Madonnetta":  300,
    "Monza":       400,
    "Parma":       200,
    "Catania":     475,
    "Firenze":     675,
    "Bologna":     650,
    "Genova":      825,
    "Palermo":    1025,
    "Torino":     1950,
    "Milano":     3375,
}

METHODS = ["no_gate", "fixed_gate", "rolling", "calibration"]
METHOD_LABELS = {
    "no_gate":     "No gate",
    "fixed_gate":  "Fixed gate",
    "rolling":     "Rolling",
    "calibration": "Calibration",
}

# ── Load all CSVs, normalise column names ─────────────────────────────────────
city_dfs = {}
for csv_path in sorted(INST_DIR.glob("*/*_results.csv")):
    city = csv_path.parent.name
    df = pd.read_csv(csv_path)

    # normalise old → new column names
    rename = {}
    if "new_best_via_ls_%" in df.columns and "ls_responsible_%" not in df.columns:
        rename["new_best_via_ls_%"] = "ls_responsible_%"
    if "new_best_via_dr_%" in df.columns and "dr_responsible_%" not in df.columns:
        rename["new_best_via_dr_%"] = "dr_responsible_%"
    if rename:
        df = df.rename(columns=rename)

    # add best_found_by_ls / best_found_without_ls if missing
    if "best_found_by_ls" not in df.columns:
        df["best_found_by_ls"]      = np.nan
        df["best_found_without_ls"] = np.nan

    df["city"] = city
    df["N"]    = N_MAP.get(city, 0)
    city_dfs[city] = df

if not city_dfs:
    print("No result CSVs found under", INST_DIR)
    sys.exit(1)

all_df = pd.concat(city_dfs.values(), ignore_index=True)
all_df = all_df.sort_values(["N", "method"]).reset_index(drop=True)

# ── Build calibration-only summary ───────────────────────────────────────────
calib_rows = []
for city, df in sorted(city_dfs.items(), key=lambda x: N_MAP.get(x[0], 0)):
    c = df[df["method"] == "calibration"].iloc[0]
    r = df[df["method"] == "rolling"].iloc[0]
    n = df[df["method"] == "no_gate"].iloc[0]
    calib_rows.append({
        "N":                   N_MAP.get(city, 0),
        "City":                city,
        "no_gate_score":       round(n["best_score"], 3),
        "rolling_score":       round(r["best_score"], 3),
        "calib_score":         round(c["best_score"], 3),
        "calib_vs_rolling":    round(c["best_score"] - r["best_score"], 3),
        "calib_vs_nogate":     round(c["best_score"] - n["best_score"], 3),
        "skip_%":              c["skip_%"],
        "ls_helped_%":         c["ls_helped_%"],
        "mean_ls_improv":      round(c["mean_ls_improv"], 5),
        "best_found_by_ls":    c["best_found_by_ls"],
        "best_found_without_ls": c["best_found_without_ls"],
        "ls_responsible_%":    c["ls_responsible_%"],
        "dr_responsible_%":    c["dr_responsible_%"],
    })
calib_df = pd.DataFrame(calib_rows).sort_values("N").reset_index(drop=True)

# ── Colour helpers ────────────────────────────────────────────────────────────
def hex_fill(hex_str):
    return PatternFill("solid", fgColor=hex_str.lstrip("#"))

GREEN_DARK  = hex_fill("1E8449")
GREEN_MED   = hex_fill("27AE60")
GREEN_LIGHT = hex_fill("A9DFBF")
YELLOW      = hex_fill("F9E79F")
RED_LIGHT   = hex_fill("F1948A")
RED_MED     = hex_fill("E74C3C")
BLUE_HDR    = hex_fill("1A5276")
TEAL_HDR    = hex_fill("117A65")
GREY_HDR    = hex_fill("566573")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
BOLD        = Font(bold=True)

thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, col_count, fill, font=None):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font or WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_data(ws, row, col_count, alt=False):
    bg = hex_fill("EBF5FB") if alt else hex_fill("FDFEFE")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = bg
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def color_delta(cell, val, pos_thresh=0.05, neg_thresh=-0.05):
    if pd.isna(val):
        return
    if val > pos_thresh:
        cell.fill = GREEN_LIGHT
        cell.font = Font(bold=True, color="1E8449")
    elif val < neg_thresh:
        cell.fill = RED_LIGHT
        cell.font = Font(bold=True, color="C0392B")
    else:
        cell.fill = YELLOW

def color_ls_pct(cell, val):
    if pd.isna(val):
        return
    if val >= 40:
        cell.fill = GREEN_LIGHT
        cell.font = Font(bold=True, color="1E8449")
    elif val >= 20:
        cell.fill = YELLOW
    else:
        cell.fill = RED_LIGHT
        cell.font = Font(bold=True, color="C0392B")

# ── Generate LS-vs-N chart PNG ────────────────────────────────────────────────
chart_path = OUT_DIR / "_ls_vs_N_chart.png"
plt.rcParams.update({"font.family":"serif","font.size":10,
                     "axes.spines.top":False,"axes.spines.right":False})

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle("LS vs D&R contribution to incumbent updates (calibration gate)",
             fontsize=12, y=1.02)

Ns   = calib_df["N"].tolist()
ls_r = calib_df["ls_responsible_%"].tolist()
dr_r = calib_df["dr_responsible_%"].tolist()
cvr  = calib_df["calib_vs_rolling"].tolist()
cities = calib_df["City"].tolist()

ax1.plot(Ns, ls_r, "o-", color="#8E44AD", linewidth=2, markersize=8, label="LS responsible %")
ax1.plot(Ns, dr_r, "s-", color="#F39C12", linewidth=2, markersize=8, label="D&R responsible %")
ax1.axhline(50, color="#999", linewidth=1, linestyle="--", alpha=0.5)
ax1.set_xscale("log")
ax1.set_xlabel("N (log scale)", fontsize=11)
ax1.set_ylabel("% of incumbent updates", fontsize=11)
ax1.set_title("Who drives new best solutions?", fontsize=10, pad=6)
ax1.legend(fontsize=9, frameon=False)
ax1.grid(axis="y", alpha=0.2, linestyle="--")
for n, ls, city in zip(Ns, ls_r, cities):
    ax1.annotate(city, (n, ls), textcoords="offset points",
                 xytext=(0, 7), ha="center", fontsize=7.5, color="#8E44AD")

bar_colors = ["#27AE60" if v > 0.05 else ("#F39C12" if abs(v) <= 0.05 else "#E74C3C")
              for v in cvr]
ax2.bar(range(len(Ns)), cvr, color=bar_colors, edgecolor="white")
ax2.axhline(0, color="#555", linewidth=1.2)
ax2.set_xticks(range(len(Ns)))
ax2.set_xticklabels([f"{c}\nN={n}" for c, n in zip(cities, Ns)], fontsize=8)
ax2.set_ylabel("Calibration - Rolling score", fontsize=11)
ax2.set_title("Calibration gain over rolling gate", fontsize=10, pad=6)
ax2.grid(axis="y", alpha=0.2, linestyle="--")
for xi, v in enumerate(cvr):
    ax2.text(xi, v + (0.05 if v >= 0 else -0.3),
             f"{v:+.2f}", ha="center", fontsize=8, fontweight="bold",
             color=bar_colors[xi])

plt.tight_layout()
fig.savefig(chart_path, dpi=160, bbox_inches="tight", facecolor="white")
plt.close(fig)
plt.rcdefaults()

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Calibration Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Calibration Summary"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

title_cell = ws1.cell(row=1, column=1,
    value="Gate LS Diagnostic — Calibration Gate Summary (seed=11, alpha=0.01, floor=0.90)")
title_cell.font = Font(size=13, bold=True, color="1A5276")
title_cell.alignment = Alignment(horizontal="left")
ws1.merge_cells("A1:N1")

cols1 = [
    "N", "City",
    "No-gate score", "Rolling score", "Calib score",
    "Calib vs Rolling", "Calib vs No-gate",
    "Skip %",
    "LS helped %", "Mean LS improv",
    "best_found_by_ls", "best_found_without_ls",
    "LS responsible %", "D&R responsible %",
]
for ci, h in enumerate(cols1, 1):
    ws1.cell(row=2, column=ci, value=h)
style_header(ws1, 2, len(cols1), BLUE_HDR)

for ri, row in calib_df.iterrows():
    r = ri + 3
    vals = [
        row["N"], row["City"],
        row["no_gate_score"], row["rolling_score"], row["calib_score"],
        row["calib_vs_rolling"], row["calib_vs_nogate"],
        row["skip_%"],
        row["ls_helped_%"], row["mean_ls_improv"],
        row["best_found_by_ls"], row["best_found_without_ls"],
        row["ls_responsible_%"], row["dr_responsible_%"],
    ]
    style_data(ws1, r, len(cols1), alt=(ri % 2 == 0))
    for ci, v in enumerate(vals, 1):
        ws1.cell(row=r, column=ci).value = v if not (isinstance(v, float) and np.isnan(v)) else "N/A"

    color_delta(ws1.cell(row=r, column=6), row["calib_vs_rolling"])   # vs rolling
    color_delta(ws1.cell(row=r, column=7), row["calib_vs_nogate"])    # vs no-gate
    color_ls_pct(ws1.cell(row=r, column=13), row["ls_responsible_%"]) # LS responsible

col_widths1 = [6, 13, 14, 13, 12, 15, 15, 8, 12, 14, 17, 20, 16, 16]
for ci, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# embed chart
if chart_path.exists():
    img = XLImage(str(chart_path))
    img.width, img.height = 900, 330
    ws1.add_image(img, f"A{len(calib_df)+5}")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — All Methods Detail
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Methods Detail")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.cell(row=1, column=1,
    value="All Methods — LS Diagnostic Detail (every city x method)").font = Font(
    size=13, bold=True, color="117A65")
ws2.merge_cells("A1:N1")

cols2 = [
    "N", "City", "Method",
    "Best score", "Gain vs no-gate", "Delta vs rolling",
    "Skip %", "Full evals",
    "LS helped %", "Mean LS improv",
    "best_found_by_ls", "best_found_without_ls",
    "LS responsible %", "D&R responsible %",
]
for ci, h in enumerate(cols2, 1):
    ws2.cell(row=2, column=ci, value=h)
style_header(ws2, 2, len(cols2), TEAL_HDR)

METHOD_ORDER = {"no_gate": 0, "fixed_gate": 1, "rolling": 2, "calibration": 3}
detail_df = all_df.copy()
detail_df["_mo"] = detail_df["method"].map(METHOD_ORDER)
detail_df = detail_df.sort_values(["N", "_mo"]).reset_index(drop=True)

prev_city = None
alt = False
for ri, row in detail_df.iterrows():
    r = ri + 3
    if row["city"] != prev_city:
        alt = not alt
        prev_city = row["city"]

    style_data(ws2, r, len(cols2), alt=alt)

    # method label colouring
    method_colors = {
        "no_gate":     "2471A3",
        "fixed_gate":  "C0392B",
        "rolling":     "27AE60",
        "calibration": "E67E22",
    }
    mcell = ws2.cell(row=r, column=3, value=METHOD_LABELS.get(row["method"], row["method"]))
    mcell.font = Font(bold=True, color=method_colors.get(row["method"], "000000"))

    ls_r_val = row.get("ls_responsible_%", np.nan)
    dr_r_val = row.get("dr_responsible_%", np.nan)

    vals = [
        row["N"], row["city"], None,  # method already written
        round(row["best_score"], 3),
        round(row["gain_vs_nogate"], 3),
        round(row["delta_vs_rolling"], 3),
        row["skip_%"], row["full_evals"],
        row["ls_helped_%"], round(row["mean_ls_improv"], 5),
        row["best_found_by_ls"] if not pd.isna(row.get("best_found_by_ls", np.nan)) else "N/A",
        row["best_found_without_ls"] if not pd.isna(row.get("best_found_without_ls", np.nan)) else "N/A",
        ls_r_val if not pd.isna(ls_r_val) else "N/A",
        dr_r_val if not pd.isna(dr_r_val) else "N/A",
    ]
    for ci, v in enumerate(vals, 1):
        if ci == 3:
            continue
        ws2.cell(row=r, column=ci).value = v

    color_delta(ws2.cell(row=r, column=5), row["gain_vs_nogate"])
    color_delta(ws2.cell(row=r, column=6), row["delta_vs_rolling"])
    if row["method"] == "calibration" and not pd.isna(ls_r_val):
        color_ls_pct(ws2.cell(row=r, column=13), ls_r_val)

col_widths2 = [6, 13, 13, 12, 15, 15, 8, 10, 12, 14, 17, 20, 16, 16]
for ci, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# Sheets 3+ — one per city
# ═══════════════════════════════════════════════════════════════════════════════
for city, df in sorted(city_dfs.items(), key=lambda x: N_MAP.get(x[0], 0)):
    ws = wb.create_sheet(city)
    ws.sheet_view.showGridLines = False

    N_val = N_MAP.get(city, "?")
    ws.cell(row=1, column=1,
        value=f"{city}  (N={N_val})  —  LS Diagnostic").font = Font(
        size=12, bold=True, color="1A5276")
    ws.merge_cells("A1:N1")

    cols = [
        "Method", "Best score", "Gain vs no-gate", "Delta vs rolling",
        "Skip %", "Full evals",
        "LS helped %", "Mean LS improv",
        "best_found_by_ls", "best_found_without_ls",
        "LS responsible %", "D&R responsible %",
    ]
    for ci, h in enumerate(cols, 1):
        ws.cell(row=2, column=ci, value=h)
    style_header(ws, 2, len(cols), GREY_HDR)

    df_s = df.copy()
    df_s["_mo"] = df_s["method"].map(METHOD_ORDER)
    df_s = df_s.sort_values("_mo").reset_index(drop=True)

    ng_score = df_s[df_s["method"] == "no_gate"]["best_score"].iloc[0]
    ro_score = df_s[df_s["method"] == "rolling"]["best_score"].iloc[0]

    for ri, row in df_s.iterrows():
        r = ri + 3
        style_data(ws, r, len(cols), alt=(ri % 2 == 0))

        ls_r_v = row.get("ls_responsible_%", np.nan)
        dr_r_v = row.get("dr_responsible_%", np.nan)
        bfl    = row.get("best_found_by_ls", np.nan)
        bfwl   = row.get("best_found_without_ls", np.nan)

        vals = [
            METHOD_LABELS.get(row["method"], row["method"]),
            round(row["best_score"], 3),
            round(row["best_score"] - ng_score, 3),
            round(row["best_score"] - ro_score, 3),
            row["skip_%"], row["full_evals"],
            row["ls_helped_%"], round(row["mean_ls_improv"], 5),
            bfl if not pd.isna(bfl) else "N/A",
            bfwl if not pd.isna(bfwl) else "N/A",
            ls_r_v if not pd.isna(ls_r_v) else "N/A",
            dr_r_v if not pd.isna(dr_r_v) else "N/A",
        ]
        mcell = ws.cell(row=r, column=1, value=vals[0])
        mcell.font = Font(bold=True, color=method_colors.get(row["method"], "000000"))
        for ci, v in enumerate(vals[1:], 2):
            ws.cell(row=r, column=ci).value = v

        color_delta(ws.cell(row=r, column=3), row["best_score"] - ng_score)
        color_delta(ws.cell(row=r, column=4), row["best_score"] - ro_score)
        if not pd.isna(ls_r_v):
            color_ls_pct(ws.cell(row=r, column=11), ls_r_v)

    col_widths = [14, 12, 15, 15, 8, 10, 12, 14, 17, 20, 16, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # embed plot if it exists
    plot_png = INST_DIR / city / f"{city}_results.png"
    if plot_png.exists():
        img = XLImage(str(plot_png))
        img.width, img.height = 900, 280
        ws.add_image(img, f"A{len(df_s)+5}")

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = OUT_DIR / args.out
wb.save(out_path)
print(f"\nExcel -> {out_path}")
print(f"Sheets: Calibration Summary, All Methods Detail, + {len(city_dfs)} city tabs")
inst_list = ", ".join(f"{c}(N={N_MAP.get(c,'?')})" for c in sorted(city_dfs, key=lambda x: N_MAP.get(x, 0)))
print(f"Instances: {inst_list}")
print("\nDone.")
