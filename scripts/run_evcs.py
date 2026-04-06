# =========================
# 0) IMPORTS & PATHS
# =========================
from pathlib import Path
import time
import pandas as pd
import sys
import numpy as np
import argparse
from datetime import datetime

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
sys.path.append(str(SRC_DIR))

from evcs import (
    build_multi_period_model,
    solve_model,
    run_DR_multi
)

from evcs.methods import sync_solution_state, reassign_y_greedy_multi
from evcs.geom import build_arcs

# =========================
# ARGUMENTS
# =========================
parser = argparse.ArgumentParser()
parser.add_argument("--csv",  type=str,   default="center_146_Verona_k250.csv")
parser.add_argument("--seed", type=int,   default=11)
parser.add_argument("--D",    type=float, default=2.0)
parser.add_argument("--T",    type=int,   default=6)
parser.add_argument("--Q",    type=float, default=20.0)
args = parser.parse_args()

# =========================
# CONFIG
# =========================
DATA_DIR  = PROJECT_ROOT / "data" / "input"
RESULTS_DIR = PROJECT_ROOT / "results"
RESULTS_DIR.mkdir(exist_ok=True)

CSV_NAME              = args.csv
D                     = args.D
T                     = args.T
Q                     = args.Q
seed                  = args.seed

policy                = "closest_priority"
max_chargers_per_site = 6

# DR parameters
max_iter       = 2000     
dr_time_limit  = 1000     
batch_size     = 50
top_k_full     = 8
ls_moves       = 12
ls_frac_remove = 0.08
frac_remove    = 0.3
accept_epsilon = 1e-6
adaptive_destroy = True
destroy_modes  = ["site_swap", "local_remove", "area_destroy"]

# Exact solver
exact_time_limit = 3600    
mip_gap          = 0.001


# =========================
# INSTANCE BUILDER
# Identical to notebook build_inst_from_csv()
# =========================
def build_instance_from_csv(csv_path, T, seed, D_km):
   
    df = pd.read_csv(csv_path)

    # coordinates degrees -> km
    coords_deg = df[["Centroid_Longitude", "Centroid_Latitude"]].to_numpy(dtype=float)
    coords_km = coords_deg.copy()
    coords_km[:, 0] *= 111.0
    coords_km[:, 1] *= 111.0

    # demand from population share
    pop = df["Aggregated_Population"].to_numpy(dtype=float)
    pop = np.maximum(pop, 0.0)
    if pop.sum() <= 0:
        raise ValueError("Population column sums to zero.")
    pop_share = pop / pop.sum()

    M = coords_km.shape[0]
    base = pop_share * M   # normalized so sum(base) = M

    # seasonal demand with noise — identical to notebook
    rng = np.random.default_rng(seed)
    demand_IT = np.zeros((T, M), dtype=float)
    for t in range(T):
        season_factor = 1.0 + 0.2 * np.sin(2 * np.pi * t / T)
        noise = rng.normal(0.0, 0.05, size=M)
        demand_IT[t] = np.maximum(0.0, base * season_factor * (1.0 + noise))

    # build arcs — forbid_self=False same as notebook
    distIJ, in_range, Ji, Ij = build_arcs(
        coords_km, coords_km, D=D_km, forbid_self=False
    )

    return {
        "coords_I":  coords_km,
        "coords_J":  coords_km.copy(),
        "demand_IT": demand_IT,
        "distIJ":    distIJ,
        "in_range":  in_range,
        "Ji":        Ji,
        "Ij":        Ij,
        "M":         M,
        "N":         M,
    }


# =========================
# MAIN
# =========================
def run_single_experiment():
    global T, seed

    print("Starting EVCS experiment")
    t_global_start = time.time()

    # -------------------------
    # LOAD — identical to notebook
    # -------------------------
    inst = build_instance_from_csv(
        DATA_DIR / CSV_NAME, T=T, seed=seed, D_km=D
    )

    M         = inst["M"]
    N         = inst["N"]
    distIJ    = inst["distIJ"]
    in_range  = inst["in_range"]
    Ji        = inst["Ji"]
    Ij        = inst["Ij"]
    demand_IT = inst["demand_IT"]   # shape (T, M)
    P_T_local = [8] * T             # budget per period

    print(f"Instance: {CSV_NAME}  N={N}  T={T}  D={D}  seed={seed}  |A|={len(in_range)}")

    # -------------------------
    # DR
    # -------------------------
    print("\n Running DR...")
    t_dr_start = time.time()

    dr_out = run_DR_multi(
        inst=inst,
        policy=policy,
        P_T=P_T_local,
        Q=Q,
        D=D,
        T=T,
        max_iter=max_iter,
        dr_time_limit=dr_time_limit,
        seed=seed,
        max_chargers_per_site=max_chargers_per_site,
        frac_remove=frac_remove,
        accept_epsilon=accept_epsilon,
        adaptive_destroy=adaptive_destroy,
        destroy_modes=destroy_modes,
        batch_size=batch_size,
        top_k_full=top_k_full,
        ls_moves=ls_moves,
        ls_frac_remove=ls_frac_remove,
    )

    t_dr_end = time.time()

    dr_best = dr_out["best_obj"]

    # -------------------------
    # EXACT
    # -------------------------
    print("\n Running Exact...")
    t_exact_start = time.time()

    model = build_multi_period_model(
        M=M, N=N, T=T,
        in_range=in_range,
        Ji=Ji,
        Ij=Ij,
        demand_IT=demand_IT,
        Q=Q,
        P_T=P_T_local,
        distIJ=distIJ,
        method_name=policy,
        max_chargers_per_site=max_chargers_per_site,
    )

    res = solve_model(
        model,
        time_limit=exact_time_limit,
        mip_gap=mip_gap,
        solver_name="gurobi",
    )

    t_exact_end = time.time()

    exact_inc_raw  = getattr(res, "best_feasible_objective", None)
    exact_bound_raw = getattr(res, "best_objective_bound", None)

    if exact_inc_raw is not None:
        exact_inc_raw = float(exact_inc_raw)
    if exact_bound_raw is not None:
        exact_bound_raw = float(exact_bound_raw)

    if exact_inc_raw is not None and exact_bound_raw is not None:
        exact_gap_raw = abs(exact_inc_raw - exact_bound_raw) / max(abs(exact_inc_raw), 1e-9)
    else:
        exact_gap_raw = None

    # -------------------------
    # Evaluate EXACT with greedy assignment
    # (same evaluation method as DR uses)
    # Only run if Gurobi actually found a feasible integer solution
    # -------------------------
    # -------------------------
    # RESULTS
    # -------------------------
    t_global_end = time.time()

    # Gap = (Exact_incumbent_raw - DR_best) / Exact_incumbent_raw
    # This is the only reported gap: DR vs the MIP's true joint-optimal objective
    if exact_inc_raw is not None and exact_inc_raw > 1e-9:
        gap_abs = exact_inc_raw - dr_best
        gap_pct = 100 * gap_abs / exact_inc_raw
    else:
        gap_abs = None
        gap_pct = None

    print("\n===== RESULTS =====")
    print(f"Instance         : {CSV_NAME}")
    print(f"N, M, T          : {N}, {M}, {T}")
    print(f"Policy           : {policy}")
    print(f"seed             : {seed}")
    print(f"DR best          : {dr_best:.4f}")
    print(f"Exact incumbent  : {exact_inc_raw:.4f}" if exact_inc_raw is not None else "Exact incumbent  : N/A")
    print(f"Gap              : {gap_abs:.4f} ({gap_pct:.4f}%)" if gap_pct is not None else "Gap              : N/A")
    print(f"Time DR          : {t_dr_end - t_dr_start:.2f}s")
    print(f"Time EX          : {t_exact_end - t_exact_start:.2f}s")
    print(f"Time TOT         : {t_global_end - t_global_start:.2f}s")
    print("=====================\n")

    # EXCEL LOG — benchmark_with_SLURM.xlsx
    # =========================
    # ---- output path ----
    bench_dir  = PROJECT_ROOT / "results" / "benchmarking"
    bench_dir.mkdir(parents=True, exist_ok=True)
    excel_file = bench_dir / "benchmark_with_SLURM.xlsx"

    # ---- build row ----
    row = {
        # run identity
        "Timestamp":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Instance":             Path(CSV_NAME).stem,
        "Policy":               policy,
        "N":                    N,
        "M":                    M,
        "T":                    T,
        "seed":                 seed,
        "D_km":                 D,
        "Q":                    Q,
        "P_T":                  str(P_T_local),
        "max_chargers_per_site": max_chargers_per_site,

        # arcs
        "|A|":                  len(in_range),
        "arc_density":          round(len(in_range) / max(M * N, 1), 6),
        "total_demand":         round(float(demand_IT.sum()), 4),

        # exact
        "Exact_incumbent_raw":  exact_inc_raw,
        "Exact_bound_raw":      exact_bound_raw,
        "Exact_gap_raw":        exact_gap_raw,
        "Exact_time_s":         round(t_exact_end - t_exact_start, 2),

        # DR
        "DR_best":              round(dr_best, 4),
        "DR_iters":             len(dr_out.get("DR_trace", [])),
        "DR_time_s":            round(t_dr_end - t_dr_start, 2),
        "DR_time_limit_s":      dr_time_limit,
        "batch_size":           batch_size,
        "top_k_full":           top_k_full,
        "ls_moves":             ls_moves,
        "max_iter":             max_iter,
        "frac_remove":          frac_remove,
        "accept_epsilon":       accept_epsilon,

        # gap: DR vs MIP true joint-optimal objective (the only reported gap)
        "Gap_%":                round(gap_pct, 4) if gap_pct is not None else None,

        # timing
        "Total_time_s":         round(t_global_end - t_global_start, 2),
    }

    df_new = pd.DataFrame([row])

    # ---- append to existing or create new ----
    df_old_sheets = {}   # will hold existing trace sheets to preserve them
    if excel_file.exists():
        try:
            df_old = pd.read_excel(excel_file, sheet_name="benchmark")

            # preserve existing trace sheets
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            for sname in wb.sheetnames:
                if sname != "benchmark":
                    df_old_sheets[sname] = pd.read_excel(excel_file, sheet_name=sname)
            wb.close()

            # add any new columns that did not exist before
            for col in df_new.columns:
                if col not in df_old.columns:
                    df_old[col] = np.nan

            # add any old columns missing in new row
            for col in df_old.columns:
                if col not in df_new.columns:
                    df_new[col] = np.nan

            # keep column order consistent
            df_new = df_new[df_old.columns]
            df_all = pd.concat([df_old, df_new], ignore_index=True)

        except Exception as e:
            print(f"WARNING: Could not read existing Excel, creating new: {e}")
            df_all = df_new
    else:
        df_all = df_new

    # ---- assign trace sheet name for this row ----
    trace     = dr_out.get("DR_trace")
    has_trace = trace is not None and len(trace) > 0
    trace_sheet_name = f"t{len(df_all) - 1}" if has_trace else None
    if trace_sheet_name:
        df_all.loc[len(df_all) - 1, "Trace_sheet"] = trace_sheet_name

    # ---- save ----
    def _write_excel(path):
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            df_all.to_excel(writer, sheet_name="benchmark", index=False)
            # write preserved trace sheets
            for sname, sdf in df_old_sheets.items():
                sdf.to_excel(writer, sheet_name=sname, index=False)
            # write new trace sheet
            if has_trace:
                trace.to_excel(writer, sheet_name=trace_sheet_name, index=False)

    try:
        _write_excel(excel_file)
        print(f"Excel updated -> {excel_file}")
        print(f"   Total rows: {len(df_all)}")
        if has_trace:
            print(f"   Trace sheet: {trace_sheet_name}  ({len(trace)} iterations)")
        else:
            print("   No trace data.")
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = excel_file.with_name(f"benchmark_with_SLURM_LOCKED_{stamp}.xlsx")
        _write_excel(alt)
        print(f"WARNING: File was open -- saved as: {alt}")
# =========================
# ENTRY
# =========================
if __name__ == "__main__":
    run_single_experiment()