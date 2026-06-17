"""
export_gate_results.py
Collects all gate experiment results and writes a formatted Excel workbook.
Output: results/gate/gate_results_summary.xlsx
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT     = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "results" / "gate" / "gate_results_summary_v3.xlsx"

# ── Color palette ──────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark blue  – sheet / section headers
C_NOGATE   = "D6E4F7"   # light blue – no gate rows
C_FIXED    = "FDECEA"   # light red  – fixed gate rows
C_ROLLING  = "E2F0D9"   # light green – rolling adaptive rows
C_CALIB    = "FFF2CC"   # light yellow – calibration rows
C_BEST     = "70AD47"   # green fill  – best score cell in each group
C_WORST    = "FF7B7B"   # red fill    – worst score (over-tightening)
C_SEC_HDR  = "BDD7EE"   # section sub-header
C_WARN     = "FFC000"   # orange – warning (over-tightening flagged)

# ── Row-colour map by config keyword ──────────────────────────────────────────
def row_color(label: str) -> str:
    l = label.lower()
    if "no gate" in l:       return C_NOGATE
    if "fixed" in l:         return C_FIXED
    if "rolling" in l:       return C_ROLLING
    if "calib" in l:         return C_CALIB
    if "stat" in l:          return C_ROLLING   # stat treated as rolling variant
    if "adaptive" in l:      return C_ROLLING
    return "FFFFFF"

def fill(hex_col: str):
    return PatternFill("solid", fgColor=hex_col)

def bold_font(size=10, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Experiment definitions ─────────────────────────────────────────────────────
experiments = [
    {
        "tag":   "no_floor",
        "name":  "Exp 1 – No floor (over-tightening baseline)",
        "time":  60,
        "notes": (
            "First test of adaptive gate — no floor on the threshold.\n"
            "Rolling-mean formula: gate = proxy_best / avg_ratio ≈ proxy_best,\n"
            "which grows with the incumbent → 65 % skip rate → worst score.\n"
            "Conclusion: floor is essential."
        ),
    },
    {
        "tag":   "adaptive_floor",
        "name":  "Exp 2 – Floor added (floor_frac = 0.90)",
        "time":  60,
        "notes": (
            "Floor = 0.90 × proxy_init caps the threshold so it never\n"
            "goes above the fixed-gate baseline. Skip rate drops to 26 %.\n"
            "Adaptive gate beats both baselines. Supervisor-approved result."
        ),
    },
    {
        "tag":   "normal_vs_rolling",
        "name":  "Exp 3 – Statistical gate on LS ratios (floor = 0.90)",
        "time":  90,
        "notes": (
            "Fit N(μ, σ²) to LS improvement ratios; threshold =\n"
            "proxy_best / norm.ppf(1-α, μ, σ).  α = 0.10.\n"
            "σ = 0.0027 (near-degenerate distribution — LS barely improves);\n"
            "statistical and rolling thresholds are virtually identical.\n"
            "Both tie at 286.571.  Statistical α has no practical effect here."
        ),
    },
    {
        "tag":   "calib_vs_rolling",
        "name":  "Exp 4 – Calibration gate, floor = 0.90  [slow machine]",
        "time":  90,
        "notes": (
            "Fit UE = β₀ + β₁·proxy on ALL (proxy, UE) pairs;\n"
            "threshold = (UE_best + σ_reg·z_α − β₀) / β₁.\n"
            "Floor = 0.90 always binding (calib wants ~2157, floor = 2004).\n"
            "Machine was ~3× slower this run — fewer iterations overall.\n"
            "Calib appears to win (+0.7) but partly a speed artefact."
        ),
    },
    {
        "tag":   "calib_no_floor",
        "name":  "Exp 5 – Calibration gate, NO floor",
        "time":  90,
        "notes": (
            "Floor removed for calibration — same feedback-loop pathology\n"
            "as Exp 1: threshold jumped to 2194, 80 % skip rate, worst score.\n"
            "Conclusion: calibration also requires a floor."
        ),
    },
    {
        "tag":   "floor097",
        "name":  "Exp 6 – Calibration gate, floor = 0.97  (Run 1)",
        "time":  90,
        "notes": (
            "Raised floor to 0.97 × proxy_init = 2160, just below where\n"
            "calibration naturally sits (2155–2178). Calibration now\n"
            "sometimes overrides the floor downward (more lenient) and is\n"
            "capped upward by the floor (prevents over-tightening).\n"
            "Best score 285.754 — better than rolling at same floor (285.312).\n"
            "Rolling hurt by higher floor; calibration self-regulates better."
        ),
    },
    {
        "tag":   "floor097_run2",
        "name":  "Exp 7 – Calibration gate, floor = 0.97  (Run 2 — consistency check)",
        "time":  90,
        "notes": (
            "Exact same settings as Exp 6, re-run to verify consistency.\n"
            "Calib: 285.743 (vs 285.754 in Run 1 — Δ = 0.011, near-identical).\n"
            "Rolling: 284.985 (vs 285.312 in Run 1).\n"
            "Calibration consistently beats rolling in both runs.\n"
            "Tiny calib variance confirms the regression threshold is stable."
        ),
    },
    {
        "tag":   "alpha0003_floor090",
        "name":  "Exp 8 – Calibration gate, alpha=0.003, floor = 0.90  *** BEST ***",
        "time":  90,
        "notes": (
            "alpha=0.003 → z_alpha=-2.75 → threshold ~150 proxy units lower than alpha=0.10.\n"
            "Calibration gate now oscillates FREELY around the floor (1919–2064),\n"
            "sometimes stricter (capped by floor at 2004), sometimes more lenient.\n"
            "Skip rate: 14.7% (vs rolling 26%) — more candidates evaluated.\n"
            "New bests found: 20 (vs rolling 12) — calibration discovers more.\n"
            "Best score 286.554 — beats rolling (286.000) by +0.554. NEW BEST RESULT."
        ),
    },
]

# ── Column definitions for the data sheets ───────────────────────────────────
DATA_COLS = [
    ("Configuration",         22),
    ("Gate frac",              9),
    ("Floor frac",             9),
    ("Time (s)",               8),
    ("Iterations",            10),
    ("Skips",                  8),
    ("Skip rate (%)",         11),
    ("Full evals",            10),
    ("New bests",              9),
    ("Best score",            10),
    ("Gain vs no-gate",       15),
    ("Δ vs rolling\n(same run)", 14),
]

# ── Load all CSV data ─────────────────────────────────────────────────────────
import csv

def load_csv(tag):
    p = ROOT / "results" / "gate" / tag / "adaptive_summary.csv"
    rows = []
    with open(p, newline="") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default empty sheet

# ────────────────────────────────────────────────────────────────────────────────
# SHEET 1 — MASTER SUMMARY (all experiments, one table)
# ────────────────────────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Summary (all experiments)")

# Determine floor_frac per experiment
FLOOR_MAP = {
    "no_floor":            ("—",    "—"),
    "adaptive_floor":      ("0.90", "0.90"),
    "normal_vs_rolling":   ("0.90", "0.90"),
    "calib_vs_rolling":    ("0.90", "0.90"),
    "calib_no_floor":      ("0.90", "none"),
    "floor097":            ("0.90", "0.97"),
    "floor097_run2":       ("0.90", "0.97"),
    "alpha0003_floor090":  ("0.90", "0.90"),
}

# Title row
ws_sum.merge_cells("A1:M1")
title_cell = ws_sum["A1"]
title_cell.value = "Gate Experiment Results — center_79_Monza_k400   (EVCS Project)"
title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
title_cell.fill  = fill(C_HEADER)
title_cell.alignment = center()
ws_sum.row_dimensions[1].height = 22

# Sub-title
ws_sum.merge_cells("A2:M2")
ws_sum["A2"].value = (
    "Comparing: No gate  |  Fixed gate  |  Adaptive rolling-mean  |  Adaptive calibration"
)
ws_sum["A2"].font      = Font(italic=True, size=10, color="1F4E79")
ws_sum["A2"].alignment = center()
ws_sum.row_dimensions[2].height = 16

ROW = 3

for exp in experiments:
    tag   = exp["tag"]
    csv_rows = load_csv(tag)
    gate_frac_fixed, floor_frac_str = FLOOR_MAP[tag]

    # Section header
    ws_sum.merge_cells(f"A{ROW}:M{ROW}")
    hdr = ws_sum[f"A{ROW}"]
    hdr.value     = f"  {exp['name']}   (time budget = {exp['time']} s)"
    hdr.font      = Font(bold=True, size=10, color="1F4E79")
    hdr.fill      = fill(C_SEC_HDR)
    hdr.alignment = left()
    ws_sum.row_dimensions[ROW].height = 15
    ROW += 1

    # Column headers
    headers = [c[0] for c in DATA_COLS]
    for ci, h in enumerate(headers, 1):
        c = ws_sum.cell(ROW, ci, h)
        c.font      = bold_font(9, "FFFFFF")
        c.fill      = fill(C_HEADER)
        c.alignment = center()
        c.border    = thin_border()
    ws_sum.row_dimensions[ROW].height = 28
    ROW += 1

    # Pre-compute per-experiment baselines for the two comparison columns
    scores      = [float(r["best_score"]) for r in csv_rows]
    best_s      = max(scores)
    worst_s     = min(scores)
    nogate_s    = next((float(r["best_score"]) for r in csv_rows
                        if "no gate" in r["config"].lower()), None)
    rolling_s   = next((float(r["best_score"]) for r in csv_rows
                        if "rolling" in r["config"].lower()), None)

    for r in csv_rows:
        cfg   = r["config"]
        score = float(r["best_score"])
        skip  = float(r["skip_rate_%"])

        if "calib" in cfg.lower():
            floor_str = floor_frac_str if tag != "calib_no_floor" else "none"
        elif "rolling" in cfg.lower() or "adaptive" in cfg.lower():
            floor_str = gate_frac_fixed if tag not in ("no_floor",) else "none"
        else:
            floor_str = "—"

        gain_str  = f"{score - nogate_s:+.3f}"  if nogate_s  is not None else "—"
        delta_str = f"{score - rolling_s:+.3f}" if rolling_s is not None else "—"
        # No-gate and rolling rows show "—" for their own delta
        if "no gate" in cfg.lower():
            gain_str = "baseline"
            delta_str = "—"
        if "rolling" in cfg.lower() or ("adaptive" in cfg.lower()
                                         and "calib" not in cfg.lower()
                                         and "stat"  not in cfg.lower()):
            delta_str = "0.000 (ref)"

        vals = [
            cfg, r["gate_frac"], floor_str, r["time_s"],
            r["iterations"], r["skips"], f"{skip:.1f}",
            r["full_evals"], r["new_bests"], score,
            gain_str, delta_str,
        ]
        bg = row_color(cfg)
        for ci, v in enumerate(vals, 1):
            cell = ws_sum.cell(ROW, ci, v)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = center() if ci != 1 else left()
            cell.font      = Font(size=9)

        # Best score column (col 10)
        score_cell = ws_sum.cell(ROW, 10)
        if score == best_s:
            score_cell.fill = fill(C_BEST)
            score_cell.font = bold_font(9, "FFFFFF")
        elif score == worst_s and skip > 50:
            score_cell.fill = fill(C_WORST)
            score_cell.font = bold_font(9, "FFFFFF")

        # Δ vs rolling column (col 12) — highlight positive green, zero/negative orange
        delta_cell = ws_sum.cell(ROW, 12)
        if delta_str not in ("—", "0.000 (ref)", "baseline"):
            try:
                dv = float(delta_str)
                if dv > 0.01:
                    delta_cell.fill = fill("C6EFCE")
                    delta_cell.font = bold_font(9, "375623")
                elif abs(dv) <= 0.01:
                    delta_cell.fill = fill("FFEB9C")
                    delta_cell.font = bold_font(9, "9C5700")
                else:
                    delta_cell.fill = fill("FFC7CE")
                    delta_cell.font = bold_font(9, "9C0006")
            except ValueError:
                pass

        ROW += 1

    # Notes row
    ws_sum.merge_cells(f"A{ROW}:M{ROW}")
    note_cell = ws_sum[f"A{ROW}"]
    note_cell.value     = f"  Finding: {exp['notes'].splitlines()[0]}"
    note_cell.font      = Font(italic=True, size=8, color="595959")
    note_cell.alignment = left()
    ws_sum.row_dimensions[ROW].height = 13
    ROW += 1

    # Spacer
    ROW += 1

# Column widths
col_widths = [c[1] for c in DATA_COLS]
for ci, w in enumerate(col_widths, 1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = w

ws_sum.freeze_panes = "A4"

# ────────────────────────────────────────────────────────────────────────────────
# SHEET 2 — PROGRESSION: one row per "best adaptive" result across experiments
# ────────────────────────────────────────────────────────────────────────────────
ws_prog = wb.create_sheet("Design Progression")

prog_data = [
    # (step, what changed, method, floor, skip%, best_score, gain_vs_nogate, verdict)
    ("Step 1", "Adaptive gate — no floor",
     "Rolling mean",  "—",    "65.2%", 282.658, "-2.18", "❌ Over-tightening (no-gate=284.84)"),
    ("Step 2", "Floor added (= 0.90 × proxy_init)",
     "Rolling mean",  "0.90", "26.5%", 286.000, "+1.16", "✓ Supervisor approved (no-gate=284.84)"),
    ("Step 3", "Statistical (LS ratio) α=0.10, floor=0.90",
     "Normal dist. on LS ratios", "0.90", "25.5%", 286.571, "+1.50",
     "= Tied with rolling — σ_ratio≈0, α has no effect. Fast-machine run."),
    ("Step 3b", "Rolling mean, same run as Step 3",
     "Rolling mean", "0.90", "25.7%", 286.571, "+1.50",
     "= Tied with statistical — confirms statistical adds nothing here"),
    ("Step 4", "Calibration proxy→UE regression, floor=0.90",
     "Linear regression", "0.90", "25.7%", 285.347, "+1.25",
     "Floor always binding — calib threshold never fires. Slow-machine run."),
    ("Step 5", "Calibration — floor removed",
     "Linear regression", "none", "79.9%", 282.619, "-2.22",
     "❌ Same feedback-loop over-tightening as Step 1"),
    ("Step 6", "Calibration floor=0.97 (×2 runs, consistent)",
     "Linear regression", "0.97", "55–57%", 285.749, "+0.91",
     "✓ Calib active but skip rate too high. Beats rolling at same floor."),
    ("Step 7", "Calibration α=0.003, floor=0.90  ← BEST GAIN",
     "Linear regression", "0.90", "14.7%", 286.554, "+1.72",
     "✓✓ Gate oscillates 1919–2064. 20 new bests vs 12. Highest gain over no-gate."),
]

ws_prog.merge_cells("A1:G1")
ws_prog["A1"].value     = "Gate Design Progression"
ws_prog["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
ws_prog["A1"].fill      = fill(C_HEADER)
ws_prog["A1"].alignment = center()
ws_prog.row_dimensions[1].height = 22

prog_headers = ["Step", "Design decision", "Threshold method",
                "Floor frac", "Skip rate", "Best score", "Gain vs no-gate", "Verdict / note"]
prog_widths  = [8, 36, 26, 10, 10, 11, 15, 48]

for ci, (h, w) in enumerate(zip(prog_headers, prog_widths), 1):
    cell = ws_prog.cell(2, ci, h)
    cell.font      = bold_font(10, "FFFFFF")
    cell.fill      = fill(C_HEADER)
    cell.alignment = center()
    cell.border    = thin_border()
    ws_prog.column_dimensions[get_column_letter(ci)].width = w
ws_prog.row_dimensions[2].height = 28

# Note explaining why gain vs no-gate is the right metric
ws_prog.merge_cells("A3:H3")
note = ws_prog["A3"]
note.value = (
    "  ⚠  Scores are NOT comparable across runs (machine speed varies). "
    "Use 'Gain vs no-gate' for fair cross-run comparison — "
    "no-gate score acts as the speed baseline for each run."
)
note.font      = Font(italic=True, size=9, color="7F3F00")
note.fill      = fill("FFF2CC")
note.alignment = left()
ws_prog.row_dimensions[3].height = 14

VERDICT_COLOR = {
    "❌": "FDECEA",
    "✓": "E2F0D9",
    "=": "FFF2CC",
    "*": "FFF2CC",
}

gains = [float(r[6]) for r in prog_data]
best_gain  = max(gains)
worst_gain = min(gains)

for ri, row in enumerate(prog_data, 4):
    step, decision, method, floor, skip, score, gain, verdict = row
    vals = [step, decision, method, floor, skip, score, gain, verdict]
    vc = VERDICT_COLOR.get(verdict[0], "FFFFFF")
    for ci, v in enumerate(vals, 1):
        cell = ws_prog.cell(ri, ci, v)
        cell.fill      = fill(vc)
        cell.border    = thin_border()
        cell.alignment = center() if ci not in (2, 8) else left()
        cell.font      = Font(size=10)
    ws_prog.row_dimensions[ri].height = 18
    # Highlight gain column
    gc = ws_prog.cell(ri, 7)
    gc.font = bold_font(10)
    if float(gain) == best_gain:
        gc.fill = fill(C_BEST)
        gc.font = bold_font(10, "FFFFFF")
    if float(gain) == worst_gain:
        gc.fill = fill(C_WORST)
        gc.font = bold_font(10, "FFFFFF")

ws_prog.freeze_panes = "A3"

# ────────────────────────────────────────────────────────────────────────────────
# SHEET 3 — PER-EXPERIMENT detail tabs
# ────────────────────────────────────────────────────────────────────────────────
short_names = {
    "no_floor":           "Exp1-NoFloor",
    "adaptive_floor":     "Exp2-Floor090",
    "normal_vs_rolling":  "Exp3-StatGate",
    "calib_vs_rolling":   "Exp4-CalibFloor090",
    "calib_no_floor":     "Exp5-CalibNoFloor",
    "floor097":           "Exp6-CalibFloor097-R1",
    "floor097_run2":      "Exp7-CalibFloor097-R2",
    "alpha0003_floor090": "Exp8-BEST-a0003",
}

for exp in experiments:
    tag      = exp["tag"]
    ws       = wb.create_sheet(short_names[tag])
    csv_rows = load_csv(tag)
    _, floor_frac_str = FLOOR_MAP[tag]

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"].value     = exp["name"]
    ws["A1"].font      = Font(bold=True, size=11, color="FFFFFF")
    ws["A1"].fill      = fill(C_HEADER)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 20

    # Notes
    note_lines = exp["notes"].splitlines()
    for li, line in enumerate(note_lines):
        ws.merge_cells(f"A{2+li}:K{2+li}")
        nc = ws[f"A{2+li}"]
        nc.value     = line
        nc.font      = Font(italic=True, size=9, color="404040")
        nc.alignment = left()
        ws.row_dimensions[2+li].height = 13

    HDR_ROW = 2 + len(note_lines) + 1

    # Column headers
    detail_cols = [
        ("Configuration", 26), ("Gate frac", 10), ("Floor frac", 10),
        ("Time (s)", 9), ("Iterations", 11), ("Skips", 8),
        ("Skip rate (%)", 12), ("Full evals", 11), ("New bests", 10),
        ("Best score", 11), ("Gain vs\nno-gate", 13), ("Δ vs rolling\n(same run)", 14),
    ]
    for ci, (h, w) in enumerate(detail_cols, 1):
        cell = ws.cell(HDR_ROW, ci, h)
        cell.font      = bold_font(9, "FFFFFF")
        cell.fill      = fill(C_HEADER)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 30

    scores    = [float(r["best_score"]) for r in csv_rows]
    best_s    = max(scores)
    nogate_s  = next((float(r["best_score"]) for r in csv_rows
                      if "no gate" in r["config"].lower()), None)
    rolling_s = next((float(r["best_score"]) for r in csv_rows
                      if "rolling" in r["config"].lower()), None)

    for ri, r in enumerate(csv_rows, HDR_ROW + 1):
        cfg   = r["config"]
        score = float(r["best_score"])
        skip  = float(r["skip_rate_%"])

        if "calib" in cfg.lower():
            floor_str = floor_frac_str if tag != "calib_no_floor" else "none"
        elif any(k in cfg.lower() for k in ("rolling", "adaptive", "stat")):
            floor_str = floor_frac_str if tag not in ("no_floor",) else "none"
        else:
            floor_str = "—"

        gain_str  = f"{score - nogate_s:+.3f}"  if nogate_s  is not None else "—"
        delta_str = f"{score - rolling_s:+.3f}" if rolling_s is not None else "—"
        if "no gate" in cfg.lower():
            gain_str = "baseline"; delta_str = "—"
        if "rolling" in cfg.lower() or ("adaptive" in cfg.lower()
                                         and "calib" not in cfg.lower()
                                         and "stat"  not in cfg.lower()):
            delta_str = "0.000 (ref)"

        vals = [cfg, r["gate_frac"], floor_str, r["time_s"],
                r["iterations"], r["skips"], f"{skip:.1f}",
                r["full_evals"], r["new_bests"], score, gain_str, delta_str]

        bg = row_color(cfg)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = center() if ci != 1 else left()
            cell.font      = Font(size=10)

        sc = ws.cell(ri, 10)
        if score == best_s:
            sc.fill = fill(C_BEST)
            sc.font = bold_font(10, "FFFFFF")
        if skip > 60:
            ws.cell(ri, 7).fill = fill(C_WARN)

        # Δ vs rolling colour coding
        dc = ws.cell(ri, 12)
        if delta_str not in ("—", "0.000 (ref)", "baseline"):
            try:
                dv = float(delta_str)
                if dv > 0.01:
                    dc.fill = fill("C6EFCE"); dc.font = bold_font(10, "375623")
                elif abs(dv) <= 0.01:
                    dc.fill = fill("FFEB9C"); dc.font = bold_font(10, "9C5700")
                else:
                    dc.fill = fill("FFC7CE"); dc.font = bold_font(10, "9C0006")
            except ValueError:
                pass

    ws.freeze_panes = f"A{HDR_ROW+1}"

    # ── Embed plots ───────────────────────────────────────────────────────────
    plots = [
        ("adaptive_trajectory.png",  "Score vs Time (trajectory)"),
        ("adaptive_convergence.png", "Skip rate & p_explore convergence"),
        ("adaptive_ls_ratios.png",   "LS improvement ratio distribution & threshold"),
    ]
    plot_row = HDR_ROW + len(csv_rows) + 3   # start two rows below the data table
    exp_dir  = ROOT / "results" / "gate" / tag

    for plot_file, plot_label in plots:
        img_path = exp_dir / plot_file
        if not img_path.exists():
            continue
        # Label above the image
        ws.merge_cells(f"A{plot_row}:K{plot_row}")
        lc = ws[f"A{plot_row}"]
        lc.value     = plot_label
        lc.font      = bold_font(10, "1F4E79")
        lc.alignment = left()
        ws.row_dimensions[plot_row].height = 15
        plot_row += 1

        img = XLImage(str(img_path))
        # Scale to fit nicely (target ~900px wide at 96dpi ≈ col A anchor)
        img.width  = int(img.width  * 0.55)
        img.height = int(img.height * 0.55)
        ws.add_image(img, f"A{plot_row}")

        # Advance past the image (approximate row height needed)
        rows_needed = max(1, int(img.height / 15) + 2)
        for r in range(plot_row, plot_row + rows_needed):
            ws.row_dimensions[r].height = 15
        plot_row += rows_needed + 1

# ────────────────────────────────────────────────────────────────────────────────
# ALPHA SWEEP SHEET
# ────────────────────────────────────────────────────────────────────────────────
import csv as _csv

sweep_csv = ROOT / "results" / "gate" / "alpha_sweep" / "run1" / "alpha_sweep_results.csv"
if sweep_csv.exists():
    ws_sw = wb.create_sheet("Alpha Sweep")

    # Title
    ws_sw.merge_cells("A1:J1")
    ws_sw["A1"].value     = "Alpha Sweep — Calibration Gate  (floor=0.90, time=90s, center_79_Monza_k400)"
    ws_sw["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
    ws_sw["A1"].fill      = fill(C_HEADER)
    ws_sw["A1"].alignment = center()
    ws_sw.row_dimensions[1].height = 22

    # Warning note
    ws_sw.merge_cells("A2:J2")
    ws_sw["A2"].value = (
        "  All configs run sequentially on the same machine — scores are directly comparable within this sheet."
    )
    ws_sw["A2"].font      = Font(italic=True, size=9, color="7F3F00")
    ws_sw["A2"].fill      = fill("FFF2CC")
    ws_sw["A2"].alignment = left()
    ws_sw.row_dimensions[2].height = 13

    # Column headers
    sw_cols = [
        ("Config", 24), ("α value", 9), ("Iterations", 11), ("Skip rate (%)", 13),
        ("Full evals", 11), ("New bests", 10), ("Best score", 11),
        ("Gain vs\nno-gate", 13), ("Δ vs rolling\n(same run)", 14), ("Verdict", 30),
    ]
    for ci, (h, w) in enumerate(sw_cols, 1):
        cell = ws_sw.cell(3, ci, h)
        cell.font      = bold_font(10, "FFFFFF")
        cell.fill      = fill(C_HEADER)
        cell.alignment = center()
        cell.border    = thin_border()
        ws_sw.column_dimensions[get_column_letter(ci)].width = w
    ws_sw.row_dimensions[3].height = 30

    sw_rows = []
    with open(sweep_csv, newline="") as f:
        sw_rows = list(_csv.DictReader(f))

    # Find baselines
    nogate_score  = next((float(r["best_score"]) for r in sw_rows
                          if "no gate" in r["config"].lower()), None)
    rolling_score = next((float(r["best_score"]) for r in sw_rows
                          if "rolling" in r["config"].lower()), None)
    best_score_all = max(float(r["best_score"]) for r in sw_rows)

    for ri, r in enumerate(sw_rows, 4):
        cfg   = r["config"]
        alpha = float(r["alpha"])
        score = float(r["best_score"])
        skip  = float(r["skip_rate_%"])
        delta = float(r["delta_vs_rolling"])

        # Verdict
        if "no gate" in cfg.lower():
            verdict = "Baseline — no gate"
            bg = C_NOGATE
        elif "fixed" in cfg.lower():
            verdict = "Fixed gate baseline"
            bg = C_FIXED
        elif "rolling" in cfg.lower():
            verdict = "Rolling mean baseline (reference)"
            bg = C_ROLLING
        elif delta > 0.25:
            verdict = "✓✓ Strong improvement over rolling"
            bg = "C6EFCE"
        elif delta > 0.05:
            verdict = "✓ Beats rolling"
            bg = "E2F0D9"
        elif abs(delta) <= 0.05:
            verdict = "= Tied with rolling (floor binding)"
            bg = "FFF2CC"
        else:
            verdict = "✗ Worse than rolling"
            bg = "FDECEA"

        vals = [cfg, alpha, r["iterations"], f"{skip:.1f}",
                r["full_evals"], r["new_bests"], score,
                r["gain_vs_nogate"], r["delta_vs_rolling"], verdict]

        for ci, v in enumerate(vals, 1):
            cell = ws_sw.cell(ri, ci, v)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = center() if ci not in (1, 10) else left()
            cell.font      = Font(size=10)
        ws_sw.row_dimensions[ri].height = 16

        # Highlight best score
        sc = ws_sw.cell(ri, 7)
        if score == best_score_all:
            sc.fill = fill(C_BEST)
            sc.font = bold_font(10, "FFFFFF")

        # Delta column colour
        dc = ws_sw.cell(ri, 9)
        if "no gate" not in cfg.lower() and "fixed" not in cfg.lower() and "rolling" not in cfg.lower():
            if delta > 0.05:
                dc.fill = fill("C6EFCE"); dc.font = bold_font(10, "375623")
            elif abs(delta) <= 0.05:
                dc.fill = fill("FFEB9C"); dc.font = bold_font(10, "9C5700")
            else:
                dc.fill = fill("FFC7CE"); dc.font = bold_font(10, "9C0006")

    ws_sw.freeze_panes = "A4"

    # Embed sweep plots
    plot_row_sw = len(sw_rows) + 6
    sweep_plots = [
        ("alpha_sweep_plot.png",        "Alpha vs score / delta vs rolling / skip rate"),
        ("alpha_sweep_trajectories.png","Score trajectories for all α values"),
    ]
    sweep_dir = ROOT / "results" / "gate" / "alpha_sweep" / "run1"
    for pf, pl in sweep_plots:
        img_path = sweep_dir / pf
        if not img_path.exists():
            continue
        ws_sw.merge_cells(f"A{plot_row_sw}:J{plot_row_sw}")
        lc = ws_sw[f"A{plot_row_sw}"]
        lc.value = pl; lc.font = bold_font(10, "1F4E79"); lc.alignment = left()
        ws_sw.row_dimensions[plot_row_sw].height = 15
        plot_row_sw += 1
        img = XLImage(str(img_path))
        img.width  = int(img.width  * 0.55)
        img.height = int(img.height * 0.55)
        ws_sw.add_image(img, f"A{plot_row_sw}")
        rows_needed = max(1, int(img.height / 15) + 2)
        for r in range(plot_row_sw, plot_row_sw + rows_needed):
            ws_sw.row_dimensions[r].height = 15
        plot_row_sw += rows_needed + 2

# ── Save ──────────────────────────────────────────────────────────────────────
OUT_PATH = ROOT / "results" / "gate" / "gate_results_summary_v4.xlsx"
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Saved -> {OUT_PATH}")
